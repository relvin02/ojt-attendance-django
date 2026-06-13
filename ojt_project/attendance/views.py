from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Count
from django.utils import timezone
from django.core.files.base import ContentFile
from datetime import datetime, timedelta
from time import perf_counter
import random
import secrets
import json
# import face_recognition  # Temporarily disabled due to dlib build issues on Windows
import numpy as np
import cv2
import io
from PIL import Image
import base64
from .utils import send_sms


from .models import (
    Student, Attendance, Company, Batch, Program, 
    Supervisor, LeaveRequest, AcademicStaff,
    StudentDocument
)
from .forms import (
    StudentForm, AttendanceForm, LeaveRequestForm, 
    LeaveApprovalForm, CompanyForm, ProgramForm, 
    BatchForm, DateRangeForm, StudentDocumentForm,
    StudentProfileForm,
)
from .forms import ImportStudentsForm

# Decorators
def admin_required(view_func):
    def wrapped_view(request, *args, **kwargs):
        if request.user.is_staff:
            return view_func(request, *args, **kwargs)
        return HttpResponseForbidden("You don't have permission to access this page.")
    return wrapped_view


def supervisor_required(view_func):
    def wrapped_view(request, *args, **kwargs):
        if hasattr(request.user, 'supervisor_profile'):
            return view_func(request, *args, **kwargs)
        return HttpResponseForbidden("You don't have permission to access this page.")
    return wrapped_view


def academic_staff_required(view_func):
    def wrapped_view(request, *args, **kwargs):
        if hasattr(request.user, 'academic_staff'):
            return view_func(request, *args, **kwargs)
        return HttpResponseForbidden("You don't have permission to access this page.")
    return wrapped_view


# Helper Functions
def extract_face_encoding(image_file):
    """
    Extract face encoding from an uploaded image file.
    Returns a tuple: (encoding_list, error_message) or (None, None) if face_recognition unavailable,
    or (None, error_message) if validation fails.
    
    Security checks:
    - Exactly ONE face must be detected
    - Face must be clearly visible
    - Returns validation feedback
    """
    try:
        import face_recognition
    except ImportError:
        # Face recognition not available on this server (e.g., PythonAnywhere)
        # Return (None, None) to allow student to be added without facial recognition
        return None, None
    
    try:
        # Read image file
        image = Image.open(image_file)
        image_array = np.array(image)
        
        # Convert RGBA to RGB if needed
        if len(image_array.shape) == 3 and image_array.shape[2] == 4:
            image_array = image_array[:, :, :3]
        
        # Detect faces
        face_locations = face_recognition.face_locations(image_array)
        
        # SECURITY: Exactly one face must be detected
        if len(face_locations) == 0:
            return None, "No face detected in the image. Please use a clear photo of your face."
        elif len(face_locations) > 1:
            return None, "Multiple faces detected. Please use a photo with only your face."
        
        # Get face encodings
        face_encodings = face_recognition.face_encodings(image_array, face_locations)
        
        if not face_encodings:
            return None, "Could not extract face features. Please use a clearer photo."
        
        # Return the face encoding as a list (for JSON serialization)
        return face_encodings[0].tolist(), None
        
    except Exception as e:
        error_msg = f"Error processing image: {str(e)}"
        print(error_msg)
        return None, "Failed to process the image. Please try a different photo."


# Home Page
def index(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('admin_dashboard')
        elif hasattr(request.user, 'academic_staff'):
            return redirect('academic_staff_dashboard')
        elif hasattr(request.user, 'supervisor_profile'):
            return redirect('supervisor_dashboard')
        else:
            if hasattr(request.user, 'student_profile'):
                return redirect('student_dashboard')
    return render(request, 'attendance/index.html')


# Authentication Views
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', '/')
            return redirect(next_url)
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'attendance/login.html')


def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


# Student Views
@login_required
def student_dashboard(request):
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Student profile not found.')
        return redirect('login')
    
    # use local time instead of UTC so "today" matches the configured TIME_ZONE
    today = timezone.localtime(timezone.now()).date()
    now = timezone.localtime(timezone.now())
    
    today_attendance = Attendance.objects.filter(
        student=student,
        date=today
    ).first()

    # If the student checked in but didn't check out and the day has ended (after 17:00), mark as absent
    now = timezone.now()
    if today_attendance and today_attendance.check_in_time and not today_attendance.check_out_time:
        end_of_day = timezone.make_aware(datetime.combine(today, datetime.strptime('17:00', '%H:%M').time()))
        if now > end_of_day:
                today_attendance.status = 'absent'
                today_attendance.save()
    recent_attendances = Attendance.objects.filter(
        student=student
    ).order_by('-date')[:10]
    
    pending_leave_requests = LeaveRequest.objects.filter(
        student=student,
        status='pending'
    )
    
    # Calculate hours
    today_attendances = Attendance.objects.filter(
        student=student,
        date=today
    )
    total_hours_today = 0
    for att in today_attendances:
        hours = att.get_hours_worked()
        total_hours_today += hours or 0
    
    # This week hours (last 7 days)
    week_start = today - timedelta(days=7)
    week_attendances = Attendance.objects.filter(
        student=student,
        date__gte=week_start,
        date__lte=today
    )
    total_hours_week = sum((att.get_hours_worked() or 0) for att in week_attendances)
    
    # This month hours
    month_start = today.replace(day=1)
    month_attendances = Attendance.objects.filter(
        student=student,
        date__gte=month_start,
        date__lte=today
    )
    total_hours_month = sum((att.get_hours_worked() or 0) for att in month_attendances)
    
    attendance_stats = {
        'present': Attendance.objects.filter(student=student, status='present').count(),
        'late': Attendance.objects.filter(student=student, status='late').count(),
        'absent': Attendance.objects.filter(student=student, status='absent').count(),
        'excused': Attendance.objects.filter(student=student, status='excused').count(),
    }
    
    # Format hours
    def format_hours(hours):
        h = int(hours)
        m = int((hours - h) * 60)
        return f"{h}h {m}m"
    
    context = {
        'student': student,
        'today_attendance': today_attendance,
        'recent_attendances': recent_attendances,
        'pending_leave_requests': pending_leave_requests,
        'attendance_stats': attendance_stats,
        'today_date': today,
        'current_time': now,
        'total_hours_today': format_hours(total_hours_today),
        'total_hours_week': format_hours(total_hours_week),
        'total_hours_month': format_hours(total_hours_month),
    }
    return render(request, 'attendance/student_dashboard.html', context)


@login_required
def student_documents(request):
    """Allow a student to upload and view their own documents."""
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, 'Student profile not found.')
        return redirect('login')
    if request.method == 'POST':
        form = StudentDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.student = student
            doc.save()
            messages.success(request, 'File uploaded successfully.')
            return redirect('student_documents')
    else:
        form = StudentDocumentForm()
    documents = StudentDocument.objects.filter(student=student)
    return render(request, 'attendance/student_documents.html', {
        'form': form,
        'documents': documents,
    })


@login_required
@academic_staff_required
def academic_student_documents(request):
    """Academic staff (dean/head) can view documents submitted by students in their program."""
    academic_staff = request.user.academic_staff
    program = academic_staff.program
    documents = StudentDocument.objects.filter(student__batch__program=program)
    return render(request, 'attendance/academic_student_documents.html', {
        'documents': documents,
        'program': program,
    })

@login_required
def check_in_out(request):
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student not found'}, status=403)
    
    if request.method == 'POST':
        request_started_at = perf_counter()

        def elapsed_ms():
            return int((perf_counter() - request_started_at) * 1000)

        data = json.loads(request.body)
        image_data = data.get('image')
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        action = data.get('action')  # 'check_in' or 'check_out'
        company_id = data.get('company_id')
        liveness_verified = bool(data.get('liveness_verified'))
        liveness_mouth_sequence_verified = bool(data.get('liveness_mouth_sequence_verified'))
        liveness_turn_sequence_verified = bool(data.get('liveness_turn_sequence_verified'))
        liveness_challenge_id = data.get('liveness_challenge_id')
        liveness_completed_actions = data.get('liveness_completed_actions') or []

        try:
            liveness_started_at_ms = int(data.get('liveness_started_at_ms', 0))
            liveness_completed_at_ms = int(data.get('liveness_completed_at_ms', 0))
        except (TypeError, ValueError):
            liveness_started_at_ms = 0
            liveness_completed_at_ms = 0

        try:
            liveness_movement_score = float(data.get('liveness_movement_score', 0))
        except (TypeError, ValueError):
            liveness_movement_score = 0

        session_challenge = request.session.get('liveness_challenge') or {}
        server_challenge_id = session_challenge.get('id')
        server_action_sequence = session_challenge.get('action_sequence')
        server_expires_at = session_challenge.get('expires_at')
        server_max_duration_ms = session_challenge.get('max_duration_ms', 25000)
        now_ts = timezone.now().timestamp()

        challenge_invalid = (
            not server_challenge_id or
            server_challenge_id != liveness_challenge_id or
            not isinstance(server_action_sequence, list) or
            server_action_sequence != liveness_completed_actions or
            not server_expires_at or
            now_ts > float(server_expires_at)
        )

        duration_ms = liveness_completed_at_ms - liveness_started_at_ms
        duration_invalid = duration_ms <= 0 or duration_ms > int(server_max_duration_ms)

        if (
            not liveness_verified or
            not liveness_mouth_sequence_verified or
            not liveness_turn_sequence_verified or
            liveness_movement_score < 0.018 or
            challenge_invalid or
            duration_invalid
        ):
            return JsonResponse({
                'success': False,
                'error': 'Live face verification failed. Please complete the randomized challenge again.'
            }, status=400)

        # Consume challenge once accepted to prevent simple replay.
        request.session.pop('liveness_challenge', None)
        
        # Use student's assigned company if not provided
        if not company_id and student.company:
            company = student.company
        else:
            # Get company from request
            try:
                company = Company.objects.get(id=company_id)
            except Company.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Invalid company selected.'})
        
        # Validate that student is assigned to this company
        if student.company and student.company.id != company.id:
            return JsonResponse({
                'success': False, 
                'error': f'You are assigned to {student.company.name}. Cannot check in to a different company.'
            })
        
        # Convert base64 image to PIL Image
        try:
            image_data = image_data.split(',')[1]
            image_bytes = base64.b64decode(image_data)
            pil_image = Image.open(io.BytesIO(image_bytes))
            
            # Face recognition validation
            known_face_encoding = student.face_encoding
            if not known_face_encoding:
                return JsonResponse({
                    'success': False, 
                    'error': 'Face encoding not registered. Please update your profile.'
                })
            
            # Verify face
            try:
                import face_recognition
                
                frame = np.array(pil_image.convert('RGB'))

                # Resize recognition frame for faster processing while keeping enough detail.
                if frame.shape[1] > 480:
                    resize_ratio = 480 / frame.shape[1]
                    frame_for_recognition = cv2.resize(
                        frame,
                        (int(frame.shape[1] * resize_ratio), int(frame.shape[0] * resize_ratio)),
                        interpolation=cv2.INTER_AREA,
                    )
                else:
                    resize_ratio = 1.0
                    frame_for_recognition = frame

                face_locations = face_recognition.face_locations(frame_for_recognition)
                
                if not face_locations:
                    return JsonResponse({
                        'success': False, 
                        'error': 'No face detected. Please try again.'
                    })
                
                # SECURITY: Only one face should be detected to prevent spoofing
                if len(face_locations) > 1:
                    return JsonResponse({
                        'success': False, 
                        'error': 'Multiple faces detected. Please provide a photo with only your face.'
                    })

                if resize_ratio != 1.0:
                    face_locations = [
                        (
                            int(top / resize_ratio),
                            int(right / resize_ratio),
                            int(bottom / resize_ratio),
                            int(left / resize_ratio),
                        )
                        for top, right, bottom, left in face_locations
                    ]
                
                face_encodings = face_recognition.face_encodings(frame, face_locations)
                
                if not face_encodings:
                    return JsonResponse({
                        'success': False, 
                        'error': 'Could not extract face features. Please try again.'
                    })
                
                # Calculate face distance for strict matching
                # This is for the currently logged-in student ONLY
                face_distance = face_recognition.face_distance(
                    [np.array(known_face_encoding)], 
                    face_encodings[0]
                )[0]
                
                # Strict tolerance: 0.4 = very strict, only exact matches
                # 0.5 = strict, 0.6 = moderate (previous value)
                FACE_MATCH_TOLERANCE = 0.4
                
                if face_distance > FACE_MATCH_TOLERANCE:
                    print(f"[SECURITY] Failed face match for student {student.id_number}: distance={face_distance:.4f} (threshold={FACE_MATCH_TOLERANCE})")
                    return JsonResponse({
                        'success': False, 
                        'error': f'Face does not match your registered profile (confidence: {(1-face_distance):.0%}). Please try again with a clearer photo.'
                    })
                
                # Additional security: Log successful authentication
                print(f"[AUTH] Student {student.id_number} authenticated with face distance: {face_distance:.4f}")
                
            except ImportError:
                return JsonResponse({
                    'success': False,
                    'error': 'Face recognition is currently unavailable. Please contact support.'
                })
            
            # Get or create today's attendance
            # ensure the date is calculated in local timezone
            today = timezone.localtime(timezone.now()).date()
            attendance, created = Attendance.objects.get_or_create(
                student=student,
                company=company,
                date=today
            )
            
            # Save image
            image_file = f"attendance_{student.id}_{action}_{timezone.now().timestamp()}.jpg"
            attendance.liveness_verified = liveness_verified

            if action == 'check_in':
                attendance.check_in_time = timezone.now()
                attendance.check_in_latitude = latitude
                attendance.check_in_longitude = longitude
                attendance.check_in_image.save(image_file, ContentFile(image_bytes))

                # determine late at check-in and set was_late flag
                check_in_local = attendance.check_in_time.astimezone(timezone.get_current_timezone())
                if check_in_local.time() > datetime.strptime('08:00', '%H:%M').time():
                    attendance.was_late = True
                    attendance.status = 'late'

                attendance.save()
                # notify emergency contact
                try:
                    contact = student.emergency_contact_number
                    if contact:
                        if contact.startswith('0'):
                            contact = '+63' + contact[1:]
                        # compose name piece if contact name exists, prefixed with Mr/Mrs
                        if student.emergency_contact_name:
                            name_piece = f" Mr/Mrs {student.emergency_contact_name},"
                        else:
                            name_piece = ""
                        # convert to local timezone for display
                        check_in_local = attendance.check_in_time.astimezone(timezone.get_current_timezone())
                        send_sms(contact, (
                            f"Good morning{name_piece} {student.first_name} {student.last_name} just checked in "
                            f"at {check_in_local.strftime('%H:%M')}. Have a productive day!"
                        ))
                except Exception:
                    pass

            else:
                attendance.check_out_time = timezone.now()
                attendance.check_out_latitude = latitude
                attendance.check_out_longitude = longitude
                attendance.check_out_image.save(image_file, ContentFile(image_bytes))
                attendance.save()
                # notify emergency contact
                try:
                    contact = student.emergency_contact_number
                    if contact:
                        if contact.startswith('0'):
                            contact = '+63' + contact[1:]
                        if student.emergency_contact_name:
                            name_piece = f" Mr./Mrs. {student.emergency_contact_name},"
                        else:
                            name_piece = ""
                        check_out_local = attendance.check_out_time.astimezone(timezone.get_current_timezone())
                        send_sms(contact, (
                            f"Good afternoon{name_piece} {student.first_name} {student.last_name} checked out at "
                            f"{check_out_local.strftime('%H:%M')}. See you tomorrow!"
                        ))
                except Exception:
                    pass
                    pass

                # On check-out, compute worked hours and set final status
                try:
                    hours = attendance.get_hours_worked()
                except Exception:
                    hours = 0

                # If they were late at check-in and still worked >=8h, keep was_late True but show present
                if attendance.was_late and hours >= 8:
                    attendance.status = 'present'
                elif attendance.was_late and hours < 8:
                    # keep late as indicator; show late badge (could also mark undertime)
                    attendance.status = 'late'
                else:
                    if hours >= 8:
                        attendance.status = 'present'
                    else:
                        attendance.status = 'undertime'

                attendance.save()
            
            return JsonResponse({
                'success': True,
                'message': 'CHECK IN SUCCESSFUL' if action == 'check_in' else 'CHECK OUT SUCCESSFUL',
                'attendance_id': attendance.id,
                'processing_time_ms': elapsed_ms(),
            })
        
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    companies = Company.objects.all()
    
    # Check if already checked in today (any company) using local date
    today = timezone.localtime(timezone.now()).date()
    today_attendance = Attendance.objects.filter(
        student=student,
        date=today
    ).first()
    
    is_checked_in = today_attendance and today_attendance.check_in_time
    is_checked_out = today_attendance and today_attendance.check_out_time

    first_turn = random.choice(['left', 'right'])
    second_turn = 'right' if first_turn == 'left' else 'left'
    liveness_challenge = {
        'id': secrets.token_urlsafe(16),
        'action_sequence': ['mouth_sequence', first_turn, second_turn],
        'max_duration_ms': 25000,
        'expires_at': timezone.now().timestamp() + 60,
    }
    request.session['liveness_challenge'] = liveness_challenge
    request.session.modified = True
    
    context = {
        'companies': companies,
        'student_company': student.company,
        'is_checked_in': is_checked_in,
        'is_checked_out': is_checked_out,
        'today_attendance': today_attendance,
        'has_face_encoding': bool(student.face_encoding),
        'liveness_challenge_json': json.dumps(liveness_challenge),
    }
    return render(request, 'attendance/check_in_out.html', context)


@login_required
def attendance_history(request):
    try:
        student = request.user.student_profile
    except AttributeError:
        messages.error(request, 'Student profile not found.')
        return redirect('login')
    
    attendances = Attendance.objects.filter(student=student).order_by('-date')
    
    # Filter by date range
    form = DateRangeForm(request.GET)
    if form.is_valid() and form.cleaned_data.get('start_date'):
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        attendances = attendances.filter(date__range=[start_date, end_date])
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        attendances = attendances.filter(status=status)
    
    context = {
        'attendances': attendances,
        'form': form,
    }
    return render(request, 'attendance/attendance_history.html', context)


@login_required
def request_leave(request):
    try:
        student = request.user.student_profile
    except AttributeError:
        messages.error(request, 'Student profile not found.')
        return redirect('login')
    
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST, request.FILES)
        if form.is_valid():
            leave_request = form.save(commit=False)
            leave_request.student = student
            # proof_image (if uploaded) will be handled by the form
            leave_request.save()
            messages.success(request, 'Leave request submitted successfully.')
            return redirect('leave_history')
    else:
        form = LeaveRequestForm()
    
    context = {'form': form}
    return render(request, 'attendance/request_leave.html', context)


@login_required
def leave_history(request):
    try:
        student = request.user.student_profile
    except AttributeError:
        messages.error(request, 'Student profile not found.')
        return redirect('login')
    
    leave_requests = LeaveRequest.objects.filter(student=student).order_by('-created_at')
    
    context = {'leave_requests': leave_requests}
    return render(request, 'attendance/leave_history.html', context)


# Student Profile Views
@login_required
def student_profile(request):
    try:
        student = request.user.student_profile
    except AttributeError:
        messages.error(request, 'Student profile not found.')
        return redirect('login')
    
    if request.method == 'POST':
        form = StudentProfileForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            # Save the student first
            student = form.save()
            
            # If a new profile image was uploaded, extract face encoding
            if request.FILES.get('profile_image'):
                face_encoding, error_msg = extract_face_encoding(request.FILES['profile_image'])
                if face_encoding:
                    student.face_encoding = face_encoding
                    student.save()
                    messages.success(request, 'Profile updated successfully! Face recognition is now enabled.')
                elif error_msg:  # Only show error if there was an actual error (not just unavailable)
                    messages.error(request, f'Profile updated, but could not register face: {error_msg}')
                else:
                    messages.success(request, 'Profile updated successfully!')
            else:
                messages.success(request, 'Profile updated successfully!')
            
            return redirect('student_profile')
    else:
        form = StudentProfileForm(instance=student)
    
    context = {
        'form': form,
        'student': student,
    }
    return render(request, 'attendance/student_profile.html', context)


@login_required
@academic_staff_required
def academic_staff_edit_student(request, student_id):
    academic_staff = request.user.academic_staff
    program = academic_staff.program
    
    # Ensure student belongs to this academic staff's program
    student = get_object_or_404(Student, id=student_id, batch__program=program)
    
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            # Save the student first
            student = form.save()
            
            # If a new profile image was uploaded, extract face encoding
            if request.FILES.get('profile_image'):
                face_encoding, error_msg = extract_face_encoding(request.FILES['profile_image'])
                if face_encoding:
                    student.face_encoding = face_encoding
                    student.save()
                    messages.success(request, f'Profile for {student.first_name} {student.last_name} updated successfully! Face recognition is now enabled.')
                elif error_msg:
                    messages.error(request, f'Profile updated for {student.first_name} {student.last_name}, but could not register face: {error_msg}')
                else:
                    messages.success(request, f'Profile for {student.first_name} {student.last_name} updated successfully!')
            else:
                messages.success(request, f'Profile for {student.first_name} {student.last_name} updated successfully!')
            
            return redirect('academic_staff_manage_students')
    else:
        form = StudentForm(instance=student)
    
    context = {
        'form': form,
        'student': student,
        'program': program,
    }
    return render(request, 'attendance/academic_staff_edit_student.html', context)


@login_required
@academic_staff_required
def academic_staff_delete_student(request, student_id):
    academic_staff = request.user.academic_staff
    program = academic_staff.program
    
    # Ensure student belongs to this academic staff's program
    student = get_object_or_404(Student, id=student_id, batch__program=program)
    
    if request.method == 'POST':
        student_name = f"{student.first_name} {student.last_name}"
        student_id_number = student.id_number
        student.delete()
        messages.success(request, f'Student {student_name} (ID: {student_id_number}) has been deleted successfully.')
        return redirect('academic_staff_manage_students')
    
    context = {
        'student': student,
        'program': program,
    }
    return render(request, 'attendance/academic_staff_delete_student.html', context)


# Supervisor Views
@login_required
@supervisor_required
def supervisor_dashboard(request):
    supervisor = request.user.supervisor_profile
    company = supervisor.company
    
    today = timezone.localtime(timezone.now()).date()
    today_attendances = Attendance.objects.filter(
        company=company,
        date=today
    )
    
    total_students = Student.objects.filter(batch__company=company).count()
    present_today = today_attendances.filter(status='present').count()
    absent_today = today_attendances.filter(status='absent').count()
    late_today = today_attendances.filter(status='late').count()
    
    pending_leave_requests = LeaveRequest.objects.filter(status='pending')
    
    # Attendance status breakdown (last 30 days)
    thirty_days_ago = today - timedelta(days=30)
    recent_attendance = Attendance.objects.filter(
        company=company,
        date__gte=thirty_days_ago
    )
    
    attendance_stats = {
        'present': recent_attendance.filter(status='present').count(),
        'late': recent_attendance.filter(status='late').count(),
        'absent': recent_attendance.filter(status='absent').count(),
        'excused': recent_attendance.filter(status='excused').count(),
        'undertime': recent_attendance.filter(status='undertime').count(),
    }
    
    # Weekly attendance trends
    weekly_data = []
    for i in range(7, -1, -1):
        date = today - timedelta(days=i)
        count = Attendance.objects.filter(
            company=company,
            date=date,
            status='present'
        ).count()
        weekly_data.append({'date': date.strftime('%a'), 'count': count})
    
    context = {
        'company': company,
        'today_attendances': today_attendances,
        'total_students': total_students,
        'present_today': present_today,
        'absent_today': absent_today,
        'late_today': late_today,
        'pending_leave_requests': pending_leave_requests,
        'attendance_stats': attendance_stats,
        'weekly_data': json.dumps(weekly_data),
    }
    return render(request, 'attendance/supervisor_dashboard.html', context)


@login_required
@supervisor_required
def manage_attendance(request):
    supervisor = request.user.supervisor_profile
    company = supervisor.company
    
    attendances = Attendance.objects.filter(company=company).order_by('-date')
    
    # Filter by date
    form = DateRangeForm(request.GET)
    if form.is_valid() and form.cleaned_data.get('start_date'):
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        attendances = attendances.filter(date__range=[start_date, end_date])
    
    context = {
        'attendances': attendances,
        'form': form,
    }
    return render(request, 'attendance/manage_attendance.html', context)


@login_required
def approve_leave(request):
    # Allow supervisors, academic staff (deans), and admins to approve/reject
    user = request.user
    if hasattr(user, 'supervisor_profile'):
        supervisor = user.supervisor_profile
        pending_leaves = LeaveRequest.objects.filter(status='pending', student__company=supervisor.company)
    elif hasattr(user, 'academic_staff'):
        academic_staff = user.academic_staff
        program = academic_staff.program
        pending_leaves = LeaveRequest.objects.filter(status='pending', student__batch__program=program)
    elif user.is_staff:
        pending_leaves = LeaveRequest.objects.filter(status='pending')
    else:
        return HttpResponseForbidden("You don't have permission to access this page.")

    if request.method == 'POST':
        leave_id = request.POST.get('leave_id')
        action = request.POST.get('action')
        remarks = request.POST.get('remarks')

        leave_request = get_object_or_404(LeaveRequest, id=leave_id)
        # Ensure approver has jurisdiction over this leave
        if hasattr(user, 'supervisor_profile') and leave_request.student.company != user.supervisor_profile.company:
            return HttpResponseForbidden("You don't have permission to modify this request.")
        if hasattr(user, 'academic_staff') and leave_request.student.batch.program != user.academic_staff.program:
            return HttpResponseForbidden("You don't have permission to modify this request.")

        leave_request.status = 'approved' if action == 'approve' else 'rejected'
        leave_request.approved_by = user
        leave_request.approval_date = timezone.now()
        leave_request.remarks = remarks
        leave_request.save()

        messages.success(request, f'Absence request {action}d successfully.')
        return redirect('approve_leave')

    context = {'pending_leaves': pending_leaves}
    return render(request, 'attendance/approve_leave.html', context)


# Admin Views
@login_required
@admin_required
def admin_dashboard(request):
    total_students = Student.objects.count()
    total_programs = Program.objects.count()
    total_companies = Company.objects.count()
    total_batches = Batch.objects.count()
    
    recent_attendances = Attendance.objects.all().order_by('-created_at')[:10]
    pending_leaves = LeaveRequest.objects.filter(status='pending').count()
    
    # Attendance status breakdown (last 30 days)
    today = timezone.localtime(timezone.now()).date()
    thirty_days_ago = today - timedelta(days=30)
    recent_attendance = Attendance.objects.filter(date__gte=thirty_days_ago)
    
    attendance_stats = {
        'present': recent_attendance.filter(status='present').count(),
        'late': recent_attendance.filter(status='late').count(),
        'absent': recent_attendance.filter(status='absent').count(),
        'excused': recent_attendance.filter(status='excused').count(),
        'undertime': recent_attendance.filter(status='undertime').count(),
    }
    
    # Weekly attendance trends (last 7 days)
    weekly_data = []
    for i in range(7, -1, -1):
        date = today - timedelta(days=i)
        count = Attendance.objects.filter(date=date, status='present').count()
        weekly_data.append({'date': date.strftime('%a'), 'count': count})
    
    context = {
        'total_students': total_students,
        'total_programs': total_programs,
        'total_companies': total_companies,
        'total_batches': total_batches,
        'recent_attendances': recent_attendances,
        'pending_leaves': pending_leaves,
        'attendance_stats': attendance_stats,
        'weekly_data': json.dumps(weekly_data),
    }
    return render(request, 'attendance/admin_dashboard.html', context)


@login_required
@admin_required
def manage_students(request):
    students = Student.objects.all()
    
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save()
            messages.success(request, 'Student added successfully.')
            return redirect('manage_students')
    else:
        form = StudentForm()
    
    # Search functionality
    query = request.GET.get('q')
    if query:
        students = students.filter(
            Q(id_number__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )
    
    context = {
        'students': students,
        'form': form,
    }
    return render(request, 'attendance/manage_students.html', context)


@login_required
@admin_required
def manage_companies(request):
    companies = Company.objects.all()
    
    if request.method == 'POST':
        form = CompanyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Company added successfully.')
            return redirect('manage_companies')
    else:
        form = CompanyForm()
    
    context = {
        'companies': companies,
        'form': form,
    }
    return render(request, 'attendance/manage_companies.html', context)


@login_required
@admin_required
def manage_programs(request):
    programs = Program.objects.all()
    
    if request.method == 'POST':
        form = ProgramForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Program added successfully.')
            return redirect('manage_programs')
    else:
        form = ProgramForm()
    
    context = {
        'programs': programs,
        'form': form,
    }
    return render(request, 'attendance/manage_programs.html', context)


@login_required
@admin_required
def manage_batches(request):
    batches = Batch.objects.all()
    
    if request.method == 'POST':
        form = BatchForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Batch added successfully.')
            return redirect('manage_batches')
    else:
        form = BatchForm()
    
    context = {
        'batches': batches,
        'form': form,
    }
    return render(request, 'attendance/manage_batches.html', context)


@login_required
@admin_required
def reports(request):
    form = DateRangeForm(request.GET)
    attendances = Attendance.objects.all()
    
    if form.is_valid() and form.cleaned_data.get('start_date'):
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        attendances = attendances.filter(date__range=[start_date, end_date])
    
    # Generate statistics
    stats = {
        'total_records': attendances.count(),
        'present': attendances.filter(status='present').count(),
        'late': attendances.filter(status='late').count(),
        'absent': attendances.filter(status='absent').count(),
        'excused': attendances.filter(status='excused').count(),
    }
    
    context = {
        'attendances': attendances,
        'form': form,
        'stats': stats,
    }
    return render(request, 'attendance/reports.html', context)


# Academic Staff Views (Dean/Head)
@login_required
@academic_staff_required
def academic_staff_dashboard(request):
    academic_staff = request.user.academic_staff
    program = academic_staff.program
    
    # Get students from the academic staff's program
    batches = Batch.objects.filter(program=program)
    students = Student.objects.filter(batch__program=program)
    
    # Get today's attendance (local timezone)
    today = timezone.localtime(timezone.now()).date()
    today_attendance = Attendance.objects.filter(
        student__batch__program=program,
        date=today
    )
    
    stats = {
        'total_students': students.count(),
        'total_batches': batches.count(),
        'today_present': today_attendance.filter(status='present').count(),
        'today_absent': today_attendance.filter(status='absent').count(),
        'today_late': today_attendance.filter(status='late').count(),
    }
    # Pending absence requests for this academic staff's program
    pending_leave_requests = LeaveRequest.objects.filter(status='pending', student__batch__program=program)

    # Attendance status breakdown (last 30 days)
    thirty_days_ago = today - timedelta(days=30)
    recent_attendance = Attendance.objects.filter(
        student__batch__program=program,
        date__gte=thirty_days_ago
    )
    
    attendance_stats = {
        'present': recent_attendance.filter(status='present').count(),
        'late': recent_attendance.filter(status='late').count(),
        'absent': recent_attendance.filter(status='absent').count(),
        'excused': recent_attendance.filter(status='excused').count(),
        'undertime': recent_attendance.filter(status='undertime').count(),
    }
    
    # Weekly attendance trends
    weekly_data = []
    for i in range(7, -1, -1):
        date = today - timedelta(days=i)
        count = Attendance.objects.filter(
            student__batch__program=program,
            date=date,
            status='present'
        ).count()
        weekly_data.append({'date': date.strftime('%a'), 'count': count})

    context = {
        'academic_staff': academic_staff,
        'program': program,
        'stats': stats,
        'students': students[:10],  # Latest 10
        'batches': batches,
        'pending_leave_requests': pending_leave_requests,
        'attendance_stats': attendance_stats,
        'weekly_data': json.dumps(weekly_data),
    }
    return render(request, 'attendance/academic_staff_dashboard.html', context)


@login_required
@academic_staff_required
def academic_staff_manage_students(request):
    academic_staff = request.user.academic_staff
    program = academic_staff.program
    
    # Only show students from this program
    students = Student.objects.filter(batch__program=program)
    
    context = {
        'students': students,
        'program': program,
    }
    return render(request, 'attendance/academic_staff_manage_students.html', context)


@login_required
@academic_staff_required
def academic_staff_import_students(request):
    academic_staff = request.user.academic_staff
    program = academic_staff.program

    if request.method == 'POST':
        form = ImportStudentsForm(request.POST, request.FILES)
        if form.is_valid():
            upload = request.FILES['file']
            filename = upload.name.lower()
            successes = []
            errors = []
            created_count = 0

            try:
                # Support xlsx via openpyxl and csv via text parsing
                if filename.endswith('.xlsx') or filename.endswith('.xls'):
                    try:
                        from openpyxl import load_workbook
                    except Exception:
                        errors.append('openpyxl is not installed. Please add it to requirements.')
                        return render(request, 'attendance/academic_staff_import_students.html', {'form': form, 'errors': errors})

                    wb = load_workbook(filename=upload, read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                elif filename.endswith('.csv'):
                    import csv, io
                    text = io.TextIOWrapper(upload.file, encoding='utf-8')
                    reader = csv.reader(text)
                    rows = list(reader)
                else:
                    errors.append('Unsupported file type. Use .xlsx or .csv')
                    return render(request, 'attendance/academic_staff_import_students.html', {'form': form, 'errors': errors})

                if not rows or len(rows) < 2:
                    errors.append('File appears empty or missing header/rows.')
                    return render(request, 'attendance/academic_staff_import_students.html', {'form': form, 'errors': errors})

                header = [str(h).strip() for h in rows[0]]
                # Expected columns: id_number, first_name, last_name, email, phone, gender, date_of_birth, address, company, batch
                for idx, row in enumerate(rows[1:], start=2):
                    try:
                        row_data = {header[i]: row[i] if i < len(row) else None for i in range(len(header))}
                        id_number = str(row_data.get('id_number') or '').strip()
                        first_name = str(row_data.get('first_name') or '').strip()
                        last_name = str(row_data.get('last_name') or '').strip()
                        email = str(row_data.get('email') or '').strip()
                        phone = str(row_data.get('phone') or '').strip()
                        gender = str(row_data.get('gender') or '').strip()
                        dob = row_data.get('date_of_birth')
                        address = str(row_data.get('address') or '').strip()
                        company_name = str(row_data.get('company') or '').strip()
                        batch_name = str(row_data.get('batch') or '').strip()

                        if not id_number or not first_name or not last_name or not batch_name:
                            errors.append(f'Row {idx}: Missing required fields (id_number/first_name/last_name/batch).')
                            continue

                        # Find batch in this program
                        try:
                            batch = Batch.objects.get(name=batch_name, program=program)
                        except Batch.DoesNotExist:
                            errors.append(f'Row {idx}: Batch "{batch_name}" not found in program.')
                            continue

                        # Company optional
                        company = None
                        if company_name:
                            company = Company.objects.filter(name__iexact=company_name).first()

                        # Skip if student exists
                        if Student.objects.filter(id_number=id_number).exists():
                            errors.append(f'Row {idx}: Student with ID {id_number} already exists.')
                            continue

                        student = Student(
                            id_number=id_number,
                            first_name=first_name,
                            last_name=last_name,
                            email=email or None,
                            phone=phone or None,
                            gender=gender or None,
                            address=address or None,
                            batch=batch,
                            company=company,
                        )

                        if dob:
                            try:
                                if isinstance(dob, str):
                                    student.date_of_birth = timezone.datetime.strptime(dob, '%Y-%m-%d').date()
                                else:
                                    # openpyxl may give a date object
                                    student.date_of_birth = dob
                            except Exception:
                                pass

                        # Create user account
                        try:
                            user = User.objects.get(username=id_number)
                        except User.DoesNotExist:
                            try:
                                user = User.objects.create_user(username=id_number, email=email or '', password=id_number)
                            except Exception:
                                import uuid
                                unique_username = f"{id_number}_{uuid.uuid4().hex[:8]}"
                                user = User.objects.create_user(username=unique_username, email=email or '', password=id_number)

                        student.user = user
                        student.save()
                        created_count += 1
                        successes.append(f'Row {idx}: Created {id_number} - {first_name} {last_name}')
                    except Exception as e:
                        errors.append(f'Row {idx}: Error - {str(e)}')

            except Exception as e:
                errors.append(f'Error processing file: {str(e)}')

            return render(request, 'attendance/academic_staff_import_students.html', {
                'form': ImportStudentsForm(),
                'successes': successes,
                'errors': errors,
                'created_count': created_count,
            })
    else:
        form = ImportStudentsForm()

    return render(request, 'attendance/academic_staff_import_students.html', {'form': form, 'program': program})


@login_required
@academic_staff_required
def academic_staff_manage_attendance(request):
    from django.core.paginator import Paginator
    
    academic_staff = request.user.academic_staff
    program = academic_staff.program
    
    # Get attendance records only for this program's students
    attendances = Attendance.objects.filter(
        student__batch__program=program
    ).select_related('student', 'company').order_by('-date', '-check_in_time')
    
    form = DateRangeForm(request.GET)
    if form.is_valid() and form.cleaned_data.get('start_date'):
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date') or start_date
        attendances = attendances.filter(date__range=[start_date, end_date])
    
    # Pagination - 10 records per page
    paginator = Paginator(attendances, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'attendances': page_obj.object_list,
        'form': form,
        'program': program,
    }
    return render(request, 'attendance/academic_staff_manage_attendance.html', context)


@login_required
@academic_staff_required
def academic_staff_add_student(request):
    academic_staff = request.user.academic_staff
    program = academic_staff.program
    
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        batch_id = request.POST.get('batch')
        
        if form.is_valid() and batch_id:
            try:
                batch = Batch.objects.get(id=batch_id, program=program)
                student = form.save(commit=False)
                student.batch = batch
                
                # Create associated user if needed
                if not form.cleaned_data.get('user'):
                    id_number = form.cleaned_data.get('id_number')
                    email = form.cleaned_data.get('email')
                    
                    # Check if user already exists with this id_number
                    try:
                        user = User.objects.get(username=id_number)
                    except User.DoesNotExist:
                        # Create new user if doesn't exist
                        try:
                            user = User.objects.create_user(
                                username=id_number,
                                email=email,
                                password=id_number  # Default password
                            )
                        except Exception as e:
                            # If username still fails, generate unique username
                            import uuid
                            unique_username = f"{id_number}_{uuid.uuid4().hex[:8]}"
                            user = User.objects.create_user(
                                username=unique_username,
                                email=email,
                                password=id_number
                            )
                    
                    student.user = user
                
                student.save()
                
                # If a profile image was uploaded, extract face encoding
                if request.FILES.get('profile_image'):
                    face_encoding, error_msg = extract_face_encoding(request.FILES['profile_image'])
                    if face_encoding:
                        student.face_encoding = face_encoding
                        student.save()
                        messages.success(request, f"Student {student.first_name} {student.last_name} added successfully! Face recognition enabled.")
                    elif error_msg:
                        messages.warning(request, f"Student {student.first_name} {student.last_name} added, but face recognition not enabled: {error_msg}")
                    else:
                        messages.success(request, f"Student {student.first_name} {student.last_name} added successfully!")
                else:
                    messages.success(request, f"Student {student.first_name} {student.last_name} added successfully!")
                
                return redirect('academic_staff_manage_students')
            except Batch.DoesNotExist:
                messages.error(request, "Invalid batch selected.")
            except Exception as e:
                messages.error(request, f"Error adding student: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields and select a valid batch.")
    else:
        form = StudentForm()
    
    # Get only batches from this program
    batches = Batch.objects.filter(program=program)
    
    context = {
        'form': form,
        'batches': batches,
        'program': program,
    }
    return render(request, 'attendance/academic_staff_add_student.html', context)

# Academic Staff Reports
@login_required
@academic_staff_required
def academic_staff_reports(request):
    """Generate reports for academic staff's program"""
    from django.db.models import Q, Count, Sum, Avg
    from datetime import timedelta
    
    academic_staff = request.user.academic_staff
    program = academic_staff.program
    
    # Get filter params
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Query attendance for program
    attendances = Attendance.objects.filter(
        student__batch__program=program
    )
    
    if start_date:
        try:
            start_date_obj = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
            attendances = attendances.filter(date__gte=start_date_obj)
        except:
            start_date = ''
    
    if end_date:
        try:
            end_date_obj = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
            attendances = attendances.filter(date__lte=end_date_obj)
        except:
            end_date = ''
    
    # Get students for this program
    students = Student.objects.filter(batch__program=program)
    
    # Calculate statistics
    total_records = attendances.count()
    present_count = attendances.filter(status='present').count()
    absent_count = attendances.filter(status='absent').count()
    late_count = attendances.filter(status='late').count()
    excused_count = attendances.filter(status='excused').count()
    
    # Student performance data
    student_stats = []
    for student in students:
        student_attendance = attendances.filter(student=student)
        # get_hours_worked may return None if check-in/out missing
        hours_worked = sum((att.get_hours_worked() or 0) for att in student_attendance)
        attendance_rate = (student_attendance.filter(status='present').count() / max(student_attendance.count(), 1)) * 100 if student_attendance.exists() else 0
        
        student_stats.append({
            'student': student,
            'total_days': student_attendance.count(),
            'present': student_attendance.filter(status='present').count(),
            'absent': student_attendance.filter(status='absent').count(),
            'late': student_attendance.filter(status='late').count(),
            'hours_worked': round(hours_worked, 2),
            'attendance_rate': round(attendance_rate, 1),
        })
    
    # Company distribution
    company_stats = {}
    for attendance in attendances:
        if attendance.company:
            if attendance.company.name not in company_stats:
                company_stats[attendance.company.name] = 0
            company_stats[attendance.company.name] += 1
    
    company_list = [{'company': k, 'count': v} for k, v in company_stats.items()]
    
    # Daily breakdown (last 7 days), using local date
    today = timezone.localtime(timezone.now()).date()
    daily_stats = []
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        day_attendance = attendances.filter(date=date)
        daily_stats.append({
            'date': date.strftime('%a, %b %d'),
            'present': day_attendance.filter(status='present').count(),
            'absent': day_attendance.filter(status='absent').count(),
            'late': day_attendance.filter(status='late').count(),
        })
    
    stats = {
        'total_records': total_records,
        'present': present_count,
        'absent': absent_count,
        'late': late_count,
        'excused': excused_count,
        'total_students': students.count(),
    }
    
    context = {
        'program': program,
        'stats': stats,
        'student_stats': student_stats,
        'company_stats': company_list,
        'daily_stats': daily_stats,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'attendance/academic_staff_reports.html', context)


@login_required
@academic_staff_required
def academic_staff_print_dtr(request):
    """Print Daily Time Record (DTR) for one or multiple students for a specific month.

    Supports actions via GET param `action`:
      - print (default): render HTML for browser printing
      - export_pdf: attempt to generate PDF server-side (requires `pdfkit` + `wkhtmltopdf`), returns PDF attachment
      - email: send generated PDF to each student's email (falls back to HTML email if PDF unavailable)
    """
    academic_staff = request.user.academic_staff
    program = academic_staff.program

    # Get month/year from GET params
    month = request.GET.get('month', timezone.now().month)
    year = request.GET.get('year', timezone.now().year)

    # Accept multiple student_id parameters (multi-select) or a single id
    student_ids = request.GET.getlist('student_id') or request.GET.get('student_id')
    if isinstance(student_ids, str) and student_ids:
        # allow comma-separated string
        student_ids = [s.strip() for s in student_ids.split(',') if s.strip()]

    action = request.GET.get('action', 'print')

    try:
        month = int(month)
        year = int(year)
    except (ValueError, TypeError):
        month = timezone.now().month
        year = timezone.now().year

    # Get all students for selector
    all_students = Student.objects.filter(batch__program=program).order_by('id_number')

    # Prepare list of records (one per selected student)
    records = []

    if student_ids:
        for sid in student_ids:
            try:
                student = Student.objects.get(id=sid, batch__program=program)
            except (Student.DoesNotExist, ValueError):
                continue

            # Compute date range
            first_day = datetime(year, month, 1).date()
            if month == 12:
                last_day = datetime(year + 1, 1, 1).date() - timedelta(days=1)
            else:
                last_day = datetime(year, month + 1, 1).date() - timedelta(days=1)

            attendances = Attendance.objects.filter(
                student=student,
                date__gte=first_day,
                date__lte=last_day
            ).order_by('date')

            attendance_dict = {att.date: att for att in attendances}

            dtr_data = []
            total_hours = 0
            totals = {'present': 0, 'absent': 0, 'late': 0, 'undertime': 0}

            current_date = first_day
            while current_date <= last_day:
                day_of_week = current_date.strftime('%A')
                if current_date in attendance_dict:
                    att = attendance_dict[current_date]
                    hours = att.get_hours_worked() or 0
                    total_hours += hours

                    dtr_data.append({
                        'date': current_date,
                        'day': day_of_week,
                        'check_in': att.check_in_time.time() if att.check_in_time else None,
                        'check_out': att.check_out_time.time() if att.check_out_time else None,
                        'hours': round(hours, 2) if hours is not None else 0,
                        'status': att.status,
                        'was_late': att.was_late,
                    })

                    if att.status in totals:
                        totals[att.status] += 1
                else:
                    dtr_data.append({
                        'date': current_date,
                        'day': day_of_week,
                        'check_in': None,
                        'check_out': None,
                        'hours': 0,
                        'status': 'no_record',
                        'was_late': False,
                    })

                current_date += timedelta(days=1)

            records.append({
                'student': student,
                'dtr_data': dtr_data,
                'total_hours': round(total_hours, 2),
                'total_present': totals['present'],
                'total_absent': totals['absent'],
                'total_late': totals['late'],
                'total_undertime': totals['undertime'],
            })

    context = {
        'records': records,
        'all_students': all_students,
        'month': month,
        'year': year,
        'month_name': datetime(year, month, 1).strftime('%B %Y'),
        'program': program,
        'action': action,
    }

    # If user requested server-side PDF export or email, render HTML and try to convert to PDF
    if action in ('export_pdf', 'email') and records:
        # Render HTML content for PDF conversion
        from django.template.loader import render_to_string
        html = render_to_string('attendance/academic_staff_print_dtr.html', context, request=request)

        pdf_bytes = None
        # Try WeasyPrint first (pure-Python, easier to install in many environments),
        # then fall back to pdfkit/wkhtmltopdf if available.
        try:
            from weasyprint import HTML, CSS

            css = CSS(string='@page { size: A4; margin: 10mm }')
            pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf(stylesheets=[css])
        except Exception:
            try:
                import pdfkit

                options = {
                    'page-size': 'A4',
                    'encoding': 'UTF-8',
                    'margin-top': '10mm',
                    'margin-bottom': '10mm',
                    'margin-left': '10mm',
                    'margin-right': '10mm',
                }
                pdf_bytes = pdfkit.from_string(html, False, options=options)
            except Exception:
                pdf_bytes = None

        if action == 'export_pdf':
            if pdf_bytes:
                response = HttpResponse(pdf_bytes, content_type='application/pdf')
                safe_name = context['month_name'].replace(' ', '_')
                response['Content-Disposition'] = f'attachment; filename="DTR_{safe_name}.pdf"'
                return response
            else:
                # Fall back to returning HTML (browser can print to PDF)
                messages.warning(request, 'Server-side PDF generation is unavailable (WeasyPrint/wkhtmltopdf not installed). Showing print-friendly HTML instead.')
                return render(request, 'attendance/academic_staff_print_dtr.html', context)

        if action == 'email':
            # Send email to each selected student's email (and company email when available)
            from django.core.mail import EmailMessage
            from attendance.utils import generate_dtr_docx
            import logging as log
            
            email_logger = log.getLogger(__name__)
            sent = []
            no_email = []
            failed = []
            
            for rec in records:
                student = rec['student']
                to_addrs = [student.email] if student.email else []
                if student.company and student.company.email:
                    to_addrs.append(student.company.email)

                # Check if student has any email address
                if not to_addrs:
                    no_email.append(f"{student.id_number} ({student.first_name} {student.last_name})")
                    continue

                subject = f"DTR for {student.first_name} {student.last_name} - {context['month_name']}"
                body = 'Please find attached the Daily Time Record (DTR) for the selected month.'
                from django.conf import settings
                email = EmailMessage(subject, body, settings.DEFAULT_FROM_EMAIL, to_addrs)

                # Try to attach PDF first (fastest if available)
                if pdf_bytes:
                    email.attach(f'DTR_{student.id_number}_{context["month_name"]}.pdf', pdf_bytes, 'application/pdf')
                    email_logger.info(f"Attaching PDF for {student.id_number}")
                else:
                    # Generate DOCX as primary fallback (more readable than HTML in email clients)
                    try:
                        docx_bytes = generate_dtr_docx(rec, context['month_name'])
                        if docx_bytes:
                            email.attach(f'DTR_{student.id_number}_{context["month_name"]}.docx', docx_bytes, 'application/vnd.openxmlformats-officedocument.wordprocessingml.document')
                            email_logger.info(f"Attaching DOCX for {student.id_number}")
                        else:
                            # Last resort: HTML file
                            email.attach(f'DTR_{student.id_number}_{context["month_name"]}.html', html, 'text/html')
                            email_logger.warning(f"Attaching HTML (fallback) for {student.id_number}")
                    except Exception as e:
                        email_logger.warning(f"Error generating DOCX for {student.id_number}: {e}. Using HTML fallback.")
                        email.attach(f'DTR_{student.id_number}_{context["month_name"]}.html', html, 'text/html')

                try:
                    email.send()
                    sent.append(student.id_number)
                except Exception as e:
                    email_logger.error(f"Failed to send email to {student.id_number}: {e}")
                    failed.append(f"{student.id_number}: {str(e)}")

            # Provide detailed feedback
            if sent:
                messages.success(request, f'✓ Email sent to {len(sent)} student(s).')
            if no_email:
                messages.warning(request, f'⚠ No email on file for: {", ".join(no_email)}')
            if failed:
                messages.error(request, f'✗ Failed to send email to: {"; ".join(failed)}')
            
            if not sent and (no_email or failed):
                messages.error(request, 'No emails were sent. Please update student email addresses.')
            
            return redirect('academic_staff_print_dtr')

    # Default: render HTML for printing
    return render(request, 'attendance/academic_staff_print_dtr.html', context)